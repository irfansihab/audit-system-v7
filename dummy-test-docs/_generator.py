"""Generator dokumen dummy untuk testing audit-system-v7.

Output: ~27 file (PDF + Excel) di /sessions/.../dummy-test-docs/
Konten: realistic Indonesian content, multi-page, dengan tabel & paragraf.
"""
from pathlib import Path
from datetime import datetime, timedelta
import random

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT = Path("/sessions/trusting-blissful-heisenberg/mnt/sistem audit v7/dummy-test-docs")
OUT.mkdir(parents=True, exist_ok=True)

# Styles
styles = getSampleStyleSheet()
H1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14, alignment=TA_CENTER, spaceAfter=12, fontName="Helvetica-Bold")
H2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=11, spaceAfter=8, fontName="Helvetica-Bold")
P = ParagraphStyle("Body", parent=styles["Normal"], fontSize=10, alignment=TA_JUSTIFY, spaceAfter=6, leading=14)
PCenter = ParagraphStyle("BodyC", parent=P, alignment=TA_CENTER)

def rp(n: int) -> str:
    return f"Rp {n:,.0f}".replace(",", ".")

def build_pdf(filename: str, flowables: list):
    path = OUT / filename
    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    doc.build(flowables)
    print(f"  ✓ {filename} ({path.stat().st_size:,} bytes)")

def std_table(data, col_widths=None):
    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#d9d9d9")),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
    ]))
    return t


# ============================================================
# REVIU PENGADAAN: KAK
# ============================================================

def make_kak(nama_obyek: str, ta: int, nilai: int, sla: str, jangka: str, filename: str):
    flow = [
        Paragraph(f"KERANGKA ACUAN KERJA", H1),
        Paragraph(f"PENGADAAN {nama_obyek.upper()}", H1),
        Paragraph(f"TAHUN ANGGARAN {ta}", PCenter),
        Spacer(1, 0.5*cm),

        Paragraph("I. LATAR BELAKANG", H2),
        Paragraph(
            f"Dalam rangka mendukung transformasi digital Kementerian Komunikasi dan Digital, "
            f"diperlukan pengadaan {nama_obyek.lower()} yang andal dan sesuai dengan kebutuhan "
            f"operasional. Pengadaan ini juga mendukung pencapaian indikator kinerja utama "
            f"unit kerja di Direktorat Jenderal Aplikasi Informatika tahun {ta}. "
            f"Berdasarkan Renstra Komdigi {ta}-{ta+4}, layanan ini termasuk prioritas tinggi.", P),

        Paragraph("II. MAKSUD DAN TUJUAN", H2),
        Paragraph(
            f"Maksud pengadaan adalah memperoleh {nama_obyek.lower()} dengan spesifikasi sesuai "
            f"kebutuhan. Tujuannya: (1) meningkatkan availability layanan publik digital; "
            f"(2) memastikan kepatuhan terhadap SNI/ISO yang relevan; "
            f"(3) mendukung mitigasi risiko operasional.", P),

        Paragraph("III. RUANG LINGKUP PEKERJAAN", H2),
        Paragraph(
            f"Penyedia bertanggung jawab atas: (a) penyediaan dan instalasi; (b) integrasi dengan "
            f"sistem eksisting; (c) pelatihan operator; (d) dukungan teknis selama masa kontrak; "
            f"(e) penyerahan dokumen administrasi dan teknis.", P),

        Paragraph("IV. SPESIFIKASI TEKNIS", H2),
        std_table([
            ["No", "Komponen", "Spesifikasi Minimum"],
            ["1", "Availability", sla],
            ["2", "Lokasi", "Jakarta (DC) + Surabaya (DRC)"],
            ["3", "Sertifikasi", "ISO 27001, ISO 27017, Tier-3"],
            ["4", "Dukungan", "24x7 helpdesk dengan response time 15 menit"],
        ], col_widths=[1*cm, 4*cm, 11*cm]),
        Spacer(1, 0.4*cm),

        Paragraph("V. JANGKA WAKTU PELAKSANAAN", H2),
        Paragraph(f"Jangka waktu pelaksanaan: {jangka}.", P),

        Paragraph("VI. PERKIRAAN NILAI PENGADAAN", H2),
        Paragraph(f"Pagu anggaran: <b>{rp(nilai)}</b> (sudah termasuk PPN). Sumber dana: DIPA Komdigi TA {ta}.", P),

        Paragraph("VII. METODE PEMILIHAN", H2),
        Paragraph(
            f"Metode pemilihan: <b>Tender / E-Purchasing</b> sesuai dengan Perpres 16/2018 "
            f"jo. Perpres 12/2021 tentang Pengadaan Barang/Jasa Pemerintah.", P),

        PageBreak(),

        Paragraph("VIII. JADWAL PELAKSANAAN", H2),
        std_table([
            ["No", "Tahapan", "Bulan"],
            ["1", "Pengumuman & Pendaftaran", "Februari"],
            ["2", "Evaluasi Penawaran", "Maret"],
            ["3", "Penetapan & Kontrak", "April"],
            ["4", "Pelaksanaan & Go-live", "Mei - Desember"],
        ], col_widths=[1*cm, 9*cm, 6*cm]),
        Spacer(1, 0.4*cm),

        Paragraph("IX. KETENTUAN LAIN", H2),
        Paragraph(
            f"Service Level Agreement (SLA) {sla} berlaku selama masa kontrak. Bila SLA tidak "
            f"tercapai, penyedia dikenakan denda 0.1% per jam ketidaktersediaan. Pembayaran "
            f"bertahap mengikuti milestone yang disepakati.", P),
        Spacer(1, 1*cm),
        Paragraph("Jakarta, 15 Januari " + str(ta), PCenter),
        Paragraph("PPK Komdigi", PCenter),
        Spacer(1, 1.5*cm),
        Paragraph("(Nama PPK)", PCenter),
        Paragraph("NIP. [DIISI]", PCenter),
    ]
    build_pdf(filename, flow)


# ============================================================
# REVIU PENGADAAN: HPS
# ============================================================

def make_hps(nama_obyek: str, ta: int, items: list, vendors: list, filename: str):
    """items: [(nama, sat, vol, harga_sat), ...]
    vendors: [(nama_vendor, total_penawaran), ...]
    """
    subtotal = sum(it[2] * it[3] for it in items)
    ppn = subtotal * 0.11
    total = subtotal + ppn

    flow = [
        Paragraph(f"HARGA PERKIRAAN SENDIRI (HPS)", H1),
        Paragraph(f"PENGADAAN {nama_obyek.upper()}", H1),
        Paragraph(f"TAHUN ANGGARAN {ta}", PCenter),
        Spacer(1, 0.5*cm),

        Paragraph("A. DASAR PENYUSUNAN HPS", H2),
        Paragraph(
            f"HPS disusun berdasarkan: (1) Perpres 16/2018 tentang Pengadaan Barang/Jasa Pemerintah "
            f"dan perubahannya; (2) Standar Biaya Masukan (SBM) tahun anggaran {ta-1}; "
            f"(3) Penjajakan harga pasar via RFI dari penyedia.", P),

        Paragraph("B. RINCIAN HPS", H2),
    ]

    rows = [["No", "Uraian", "Sat", "Vol", "Harga Satuan", "Jumlah"]]
    for i, (nama, sat, vol, harga) in enumerate(items, 1):
        rows.append([str(i), nama, sat, str(vol), rp(harga), rp(vol * harga)])
    rows.append(["", "", "", "", "Subtotal", rp(subtotal)])
    rows.append(["", "", "", "", "PPN 11%", rp(int(ppn))])
    rows.append(["", "", "", "", "TOTAL", rp(int(total))])
    flow.append(std_table(rows, col_widths=[1*cm, 6*cm, 1.5*cm, 1.5*cm, 3*cm, 3*cm]))
    flow.append(Spacer(1, 0.4*cm))

    flow.append(Paragraph("C. PENJAJAKAN HARGA PASAR (RFI)", H2))
    rfi_rows = [["No", "Vendor", "Status", "Total Penawaran"]]
    for i, (v_name, v_total) in enumerate(vendors, 1):
        status = "Memberikan harga" if v_total > 0 else "Menolak partisipasi"
        rfi_rows.append([str(i), v_name, status, rp(v_total) if v_total > 0 else "—"])
    flow.append(std_table(rfi_rows, col_widths=[1*cm, 6*cm, 5*cm, 4*cm]))

    flow.append(Spacer(1, 1*cm))
    flow.append(Paragraph(f"Jakarta, 20 Januari {ta}", PCenter))
    flow.append(Paragraph("Pejabat Pengadaan Komdigi", PCenter))
    flow.append(Spacer(1, 1.5*cm))
    flow.append(Paragraph("(Nama Pejabat)", PCenter))
    flow.append(Paragraph("NIP. [DIISI]", PCenter))

    build_pdf(filename, flow)


# ============================================================
# REVIU PENGADAAN: RFI
# ============================================================

def make_rfi(vendor: str, alamat: str, kepada_obyek: str, ta: int, harga: int | None, filename: str, alasan_tolak: str = None):
    today = f"{random.randint(10,28)} Januari {ta}"
    flow = [
        Paragraph(vendor.upper(), H1),
        Paragraph(alamat, PCenter),
        Spacer(1, 0.5*cm),

        Paragraph(f"Jakarta, {today}", P),
        Spacer(1, 0.2*cm),
        Paragraph(f"Nomor: {random.randint(100,999)}/{vendor[:3].upper()}/I/{ta}", P),
        Paragraph(f"Perihal: Tanggapan Permintaan Informasi Harga (RFI) {kepada_obyek}", P),
        Spacer(1, 0.3*cm),
        Paragraph("Kepada Yth.<br/>Pejabat Pengadaan<br/>Kementerian Komunikasi dan Digital<br/>di Jakarta", P),
        Spacer(1, 0.5*cm),
        Paragraph("Dengan hormat,", P),
    ]

    if harga is None:
        # Surat penolakan
        flow.extend([
            Paragraph(
                f"Menanggapi surat permintaan informasi harga dari Kementerian Komunikasi dan Digital "
                f"perihal {kepada_obyek} tahun anggaran {ta}, dengan ini kami sampaikan bahwa "
                f"<b>{vendor} tidak dapat memberikan penawaran harga</b> untuk kegiatan tersebut.", P),
            Spacer(1, 0.2*cm),
            Paragraph(f"Alasan: {alasan_tolak or 'kapasitas tim sedang penuh untuk periode tersebut.'}", P),
            Paragraph(
                "Atas perhatiannya, kami ucapkan terima kasih. Kami tetap terbuka untuk pengadaan "
                "periode berikutnya.", P),
        ])
    else:
        flow.extend([
            Paragraph(
                f"Menanggapi surat permintaan informasi harga dari Kementerian Komunikasi dan Digital "
                f"perihal {kepada_obyek} tahun anggaran {ta}, dengan ini kami sampaikan penawaran harga "
                f"sebagai berikut:", P),
            Spacer(1, 0.3*cm),
        ])

        items = [
            ["No", "Komponen", "Harga"],
            ["1", "Layanan utama (12 bulan)", rp(int(harga * 0.85))],
            ["2", "Implementasi & integrasi", rp(int(harga * 0.10))],
            ["3", "Pelatihan & dokumentasi", rp(int(harga * 0.05))],
            ["", "TOTAL (belum termasuk PPN)", rp(harga)],
        ]
        flow.append(std_table(items, col_widths=[1*cm, 10*cm, 5*cm]))
        flow.append(Spacer(1, 0.3*cm))
        flow.append(Paragraph(
            f"Harga di atas berlaku selama 60 hari kalender. SLA layanan: 99,9% dengan dukungan 24x7. "
            f"Lokasi DC: Jakarta, DRC: Surabaya. Bersertifikat ISO 27001 dan Tier-3.", P))

    flow.extend([
        Spacer(1, 0.5*cm),
        Paragraph("Hormat kami,", P),
        Spacer(1, 1.5*cm),
        Paragraph(f"({vendor})", P),
        Paragraph("Direktur Sales", P),
    ])
    build_pdf(filename, flow)


# ============================================================
# REVIU PENGADAAN: KONTRAK
# ============================================================

def make_kontrak(no: str, nama_obyek: str, ta: int, vendor: str, nilai: int, jangka: str, filename: str):
    flow = [
        Paragraph(f"PERJANJIAN / KONTRAK", H1),
        Paragraph(f"PENGADAAN {nama_obyek.upper()}", H1),
        Paragraph(f"Nomor: {no}", PCenter),
        Spacer(1, 0.5*cm),

        Paragraph("PARA PIHAK", H2),
        Paragraph(
            f"<b>Pihak Pertama:</b> PPK Kementerian Komunikasi dan Digital, dalam hal ini "
            f"bertindak untuk dan atas nama Pemerintah RI.", P),
        Paragraph(
            f"<b>Pihak Kedua:</b> {vendor}, perusahaan yang berkedudukan di Jakarta, dalam "
            f"hal ini bertindak sebagai penyedia barang/jasa.", P),
        Spacer(1, 0.2*cm),

        Paragraph("PASAL 1 — RUANG LINGKUP", H2),
        Paragraph(
            f"Pihak Kedua menyediakan {nama_obyek.lower()} kepada Pihak Pertama sesuai dengan "
            f"Kerangka Acuan Kerja (KAK) yang menjadi bagian tidak terpisahkan dari perjanjian ini.", P),

        Paragraph("PASAL 2 — NILAI KONTRAK", H2),
        Paragraph(f"Nilai total kontrak: <b>{rp(nilai)}</b> (sudah termasuk PPN 11%).", P),

        Paragraph("PASAL 3 — JANGKA WAKTU", H2),
        Paragraph(f"Jangka waktu pelaksanaan: {jangka}.", P),

        Paragraph("PASAL 4 — SLA & DENDA", H2),
        Paragraph(
            f"Service Level Agreement: 99,9% availability per bulan. Bila SLA tidak tercapai, "
            f"Pihak Kedua dikenakan denda sebesar 0.1% dari nilai kontrak per jam ketidaktersediaan, "
            f"dengan maksimum denda kumulatif 5% dari nilai kontrak.", P),

        Paragraph("PASAL 5 — PEMBAYARAN", H2),
        Paragraph(
            f"Pembayaran dilakukan dalam 4 termin: (a) 20% setelah BAST kick-off; "
            f"(b) 30% setelah BAST implementasi; (c) 30% setelah BAST go-live; "
            f"(d) 20% setelah serah terima akhir.", P),

        Paragraph("PASAL 6 — PENYELESAIAN SENGKETA", H2),
        Paragraph(
            f"Sengketa diselesaikan secara musyawarah. Bila tidak tercapai, melalui BANI "
            f"sesuai dengan UU Arbitrase.", P),
        Spacer(1, 1*cm),

        Paragraph("Jakarta, " + f"15 Mei {ta}", PCenter),
        Spacer(1, 0.3*cm),
    ]
    # 2 column signatures
    ttd = Table([
        ["PIHAK PERTAMA", "PIHAK KEDUA"],
        ["", ""],
        ["", ""],
        ["", ""],
        ["(PPK Komdigi)", f"({vendor})"],
        ["NIP. [DIISI]", "Direktur"],
    ], colWidths=[8*cm, 8*cm])
    ttd.setStyle(TableStyle([
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ]))
    flow.append(ttd)
    build_pdf(filename, flow)


# ============================================================
# REVIU RKA-K/L: TOR
# ============================================================

def make_tor(kegiatan: str, ta: int, output: str, target: str, nilai: int, filename: str):
    flow = [
        Paragraph(f"TERM OF REFERENCE (TOR)", H1),
        Paragraph(f"KEGIATAN {kegiatan.upper()}", H1),
        Paragraph(f"TAHUN ANGGARAN {ta}", PCenter),
        Spacer(1, 0.5*cm),

        Paragraph("1. LATAR BELAKANG", H2),
        Paragraph(
            f"Kegiatan {kegiatan} merupakan salah satu agenda prioritas Kementerian Komunikasi "
            f"dan Digital tahun {ta}. Dalam rangka mencapai sasaran tersebut, diperlukan "
            f"alokasi anggaran yang memadai dan mekanisme pelaksanaan yang akuntabel.", P),

        Paragraph("2. DASAR HUKUM", H2),
        Paragraph(
            "a. UU 25/2009 tentang Pelayanan Publik;<br/>"
            "b. PP 71/2019 tentang Penyelenggaraan Sistem dan Transaksi Elektronik;<br/>"
            f"c. Perpres tentang APBN TA {ta};<br/>"
            f"d. PMK 107/2024 tentang Petunjuk Penyusunan dan Penelaahan RKA-K/L.", P),

        Paragraph("3. MAKSUD DAN TUJUAN", H2),
        Paragraph(
            f"Maksud: terlaksananya kegiatan {kegiatan} secara efektif dan efisien. "
            f"Tujuan: tercapainya output dan outcome sesuai indikator kinerja yang ditetapkan.", P),

        Paragraph("4. SASARAN, OUTPUT, DAN INDIKATOR", H2),
        std_table([
            ["Aspek", "Uraian"],
            ["Sasaran", f"Meningkatnya kapasitas operasional untuk {kegiatan.lower()}"],
            ["Output", output],
            ["Target", target],
            ["Indikator Kinerja", "Persentase capaian output ≥ 95%"],
        ], col_widths=[4*cm, 12*cm]),
        Spacer(1, 0.3*cm),

        Paragraph("5. RUANG LINGKUP", H2),
        Paragraph(
            "Lingkup kegiatan meliputi: (a) persiapan dan koordinasi; (b) pelaksanaan kegiatan inti; "
            "(c) monitoring & evaluasi; (d) pelaporan dan diseminasi hasil.", P),

        Paragraph("6. STRATEGI PELAKSANAAN", H2),
        Paragraph(
            "Kegiatan dilaksanakan dengan pendekatan kolaboratif lintas unit kerja. "
            "Penanggung jawab utama: Eselon II terkait. Tim pelaksana terdiri atas auditor, "
            "analis kebijakan, dan staf teknis.", P),

        PageBreak(),

        Paragraph("7. ANGGARAN", H2),
        Paragraph(f"Pagu anggaran kegiatan ini: <b>{rp(nilai)}</b>", P),
        Spacer(1, 0.2*cm),
        std_table([
            ["Akun", "Uraian", "Jumlah"],
            ["521211", "Belanja Bahan", rp(int(nilai*0.10))],
            ["521213", "Honorarium Output Kegiatan", rp(int(nilai*0.20))],
            ["521219", "Belanja Barang Non Operasional Lainnya", rp(int(nilai*0.15))],
            ["524111", "Belanja Perjalanan Biasa", rp(int(nilai*0.40))],
            ["522151", "Belanja Jasa Profesi", rp(int(nilai*0.15))],
            ["", "TOTAL", rp(nilai)],
        ], col_widths=[2.5*cm, 10*cm, 3.5*cm]),
        Spacer(1, 0.3*cm),

        Paragraph("8. JADWAL", H2),
        std_table([
            ["Tahap", "Bulan"],
            ["Persiapan", "Januari - Februari"],
            ["Pelaksanaan", "Maret - Oktober"],
            ["Pelaporan", "November - Desember"],
        ], col_widths=[10*cm, 6*cm]),
        Spacer(1, 1*cm),
        Paragraph(f"Jakarta, 5 Januari {ta}", PCenter),
        Paragraph("Penanggung Jawab Kegiatan", PCenter),
        Spacer(1, 1.5*cm),
        Paragraph("(Nama PJ)", PCenter),
        Paragraph("NIP. [DIISI]", PCenter),
    ]
    build_pdf(filename, flow)


# ============================================================
# REVIU RKA-K/L: RAB (Excel)
# ============================================================

def make_rab(kegiatan: str, ta: int, items: list, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "RAB"

    # Header
    ws["A1"] = f"RENCANA ANGGARAN BIAYA (RAB)"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Kegiatan: {kegiatan}"
    ws["A2"].font = Font(size=11, bold=True)
    ws.merge_cells("A2:F2")

    ws["A3"] = f"Tahun Anggaran: {ta}"
    ws.merge_cells("A3:F3")

    # Table header
    headers = ["No", "Akun", "Uraian", "Vol", "Harga Satuan", "Jumlah"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

    # Data rows
    total = 0
    for i, (akun, uraian, vol, harga) in enumerate(items, 1):
        row = i + 5
        jumlah = vol * harga
        total += jumlah
        for col, val in enumerate([str(i), akun, uraian, vol, harga, jumlah], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )
            if col in (5, 6):
                cell.number_format = '"Rp"#,##0;[Red]"-Rp"#,##0'

    # Total row
    total_row = len(items) + 6
    ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = '"Rp"#,##0;[Red]"-Rp"#,##0'

    # Column widths
    widths = [4, 10, 50, 8, 18, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+col)].width = w

    path = OUT / filename
    wb.save(str(path))
    print(f"  ✓ {filename} ({path.stat().st_size:,} bytes)")


# ============================================================
# UMUM: ST, KP, PKP
# ============================================================

def make_st(nomor: str, perihal: str, ta: int, tim: list, filename: str):
    flow = [
        Paragraph("SURAT TUGAS", H1),
        Paragraph(f"Nomor: {nomor}", PCenter),
        Spacer(1, 0.5*cm),

        Paragraph("Dasar", H2),
        Paragraph(
            "a. PP 60/2008 tentang Sistem Pengendalian Intern Pemerintah;<br/>"
            f"b. PKPT Inspektorat II TA {ta};<br/>"
            "c. Perintah Inspektur Jenderal.", P),

        Paragraph("Menugaskan", H2),
        Paragraph("Kepada nama-nama tersebut di bawah ini:", P),

        std_table(
            [["No", "Nama", "NIP", "Peran"]] + [
                [str(i+1), nm, nip, peran] for i, (nm, nip, peran) in enumerate(tim)
            ],
            col_widths=[1*cm, 6*cm, 5*cm, 4*cm]
        ),
        Spacer(1, 0.3*cm),

        Paragraph("Untuk", H2),
        Paragraph(perihal, P),

        Paragraph("Waktu", H2),
        Paragraph(f"15 hari kerja, mulai 5 Januari {ta}", P),
        Spacer(1, 1*cm),

        Paragraph(f"Jakarta, 2 Januari {ta}", PCenter),
        Paragraph("Inspektur Jenderal", PCenter),
        Spacer(1, 1.5*cm),
        Paragraph("(Nama Inspektur)", PCenter),
        Paragraph("NIP. [DIISI]", PCenter),
    ]
    build_pdf(filename, flow)


def make_kp(nomor: str, st_nomor: str, perihal: str, ta: int, filename: str):
    flow = [
        Paragraph("KARTU PENUGASAN", H1),
        Paragraph(f"Nomor: {nomor}", PCenter),
        Spacer(1, 0.5*cm),
        Paragraph(f"Dasar: Surat Tugas Nomor {st_nomor}", P),
        Paragraph(f"Perihal: {perihal}", P),
        Spacer(1, 0.3*cm),

        Paragraph("Tim Pengawasan:", H2),
        std_table([
            ["Peran", "Nama", "Jabatan Fungsional"],
            ["Pengendali Mutu", "Inspektur II", "Inspektur"],
            ["Pengendali Teknis", "Audi Wibowo, SE., M.Ak.", "Auditor Madya"],
            ["Ketua Tim", "Budi Hartono, SE., Ak.", "Auditor Muda"],
            ["Anggota Tim", "Sarah Aulia, S.Kom.", "Auditor Pertama"],
            ["Anggota Tim", "Citra Lestari, SE.", "Auditor Pertama"],
        ], col_widths=[5*cm, 6*cm, 5*cm]),
        Spacer(1, 0.5*cm),

        Paragraph("Jadwal Penugasan:", H2),
        std_table([
            ["Tahap", "Hari Ke", "Aktivitas"],
            ["Persiapan", "1-2", "Pengumpulan dokumen"],
            ["Pelaksanaan", "3-12", "Reviu substantif"],
            ["Pelaporan", "13-15", "Penyusunan LHR"],
        ], col_widths=[5*cm, 3*cm, 8*cm]),
        Spacer(1, 1*cm),
        Paragraph(f"Jakarta, 4 Januari {ta}", PCenter),
        Paragraph("Pengendali Mutu", PCenter),
    ]
    build_pdf(filename, flow)


def make_pkp(perihal: str, ta: int, sasaran: list, filename: str):
    flow = [
        Paragraph("PROGRAM KERJA PENGAWASAN (PKP)", H1),
        Paragraph(perihal, H1),
        Paragraph(f"TAHUN ANGGARAN {ta}", PCenter),
        Spacer(1, 0.5*cm),

        Paragraph("A. DASAR", H2),
        Paragraph("Surat Tugas Inspektur Jenderal Komdigi, PKPT Inspektorat II.", P),

        Paragraph("B. TUJUAN", H2),
        Paragraph(
            "Memberikan keyakinan terbatas atas kepatuhan obyek reviu terhadap peraturan "
            "yang berlaku dan kewajaran kondisi yang teridentifikasi.", P),

        Paragraph("C. RUANG LINGKUP", H2),
        Paragraph(
            "Reviu mencakup dokumen perencanaan, pelaksanaan, dan dokumen pendukung yang "
            "diserahkan oleh auditi.", P),

        Paragraph("D. SASARAN REVIU", H2),
    ]
    rows = [["ID", "Sasaran", "Langkah Kerja", "Penanggung Jawab"]]
    for sid, desc, langkah, pj in sasaran:
        rows.append([sid, desc, langkah, pj])
    flow.append(std_table(rows, col_widths=[2*cm, 5*cm, 6*cm, 3*cm]))
    flow.append(Spacer(1, 0.3*cm))

    flow.append(Paragraph("E. ALOKASI WAKTU", H2))
    flow.append(Paragraph("Total: 15 hari kerja", P))
    flow.append(Spacer(1, 1*cm))
    flow.append(Paragraph(f"Jakarta, 4 Januari {ta}", PCenter))
    flow.append(Paragraph("Pengendali Teknis", PCenter))
    flow.append(Spacer(1, 1.2*cm))
    flow.append(Paragraph("(Nama PT)", PCenter))
    flow.append(Paragraph("NIP. [DIISI]", PCenter))
    build_pdf(filename, flow)


# ============================================================
# GENERATE ALL
# ============================================================

print("="*60)
print("GENERATING DUMMY DOCUMENTS")
print(f"Output folder: {OUT}")
print("="*60)

print("\n[1] KAK (Kerangka Acuan Kerja) — 5 file")
make_kak("Layanan Cloud PSrE Induk", 2026, 5_576_851_733, "99,9%", "12 bulan",
         "KAK-Layanan-Cloud-PSrE-2026.pdf")
make_kak("Data Center & DRC Tier-3", 2026, 8_200_000_000, "99,95%", "24 bulan",
         "KAK-Data-Center-DRC-2026.pdf")
make_kak("Pelatihan Cyber Security Auditor", 2026, 450_000_000, "N/A", "6 bulan",
         "KAK-Pelatihan-Cyber-Security-2026.pdf")
make_kak("Pengembangan Aplikasi Mobile SatuData", 2026, 1_850_000_000, "99,5%", "9 bulan",
         "KAK-App-SatuData-2026.pdf")
make_kak("Penyediaan Lisensi Microsoft 365", 2026, 3_400_000_000, "99,9%", "12 bulan",
         "KAK-Lisensi-M365-2026.pdf")

print("\n[2] HPS (Harga Perkiraan Sendiri) — 4 file")
make_hps("Layanan Cloud PSrE Induk", 2026,
    items=[
        ("Compute (vCPU x 100, 12 bulan)", "Bulan", 12, 280_000_000),
        ("Storage SSD 500 TB", "Bulan", 12, 95_000_000),
        ("Network bandwidth 10 Gbps", "Bulan", 12, 45_000_000),
        ("Manajemen & dukungan 24x7", "Bulan", 12, 35_000_000),
    ],
    vendors=[
        ("PT Cyber Network Indonesia (CNI)", 6_017_600_000),
        ("PT Telkom Sigma", 0),  # menolak
        ("PT Biznet", 5_980_000_000),
        ("PT Lintasarta", 6_120_000_000),
    ],
    filename="HPS-Cloud-PSrE-2026.pdf")

make_hps("Data Center Tier-3", 2026,
    items=[
        ("Sewa rak 12 unit (24 bulan)", "Bulan", 24, 150_000_000),
        ("Power & cooling premium", "Bulan", 24, 60_000_000),
        ("Cross-connect dan link DC-DRC", "Lump sum", 1, 800_000_000),
        ("Migrasi data dari DC lama", "Lump sum", 1, 400_000_000),
    ],
    vendors=[
        ("PT DCI Indonesia", 8_450_000_000),
        ("PT NTT Indonesia", 8_690_000_000),
    ],
    filename="HPS-Data-Center-2026.pdf")

make_hps("Pelatihan Cyber Security", 2026,
    items=[
        ("Instruktur Senior (40 jam x 5 sesi)", "Sesi", 5, 60_000_000),
        ("Material & lab praktik", "Paket", 1, 80_000_000),
        ("Sertifikasi peserta (50 orang)", "Orang", 50, 1_400_000),
    ],
    vendors=[
        ("PT Multimatics", 450_000_000),
        ("PT Sapta Saka Mitra", 475_000_000),
    ],
    filename="HPS-Pelatihan-Cyber-2026.pdf")

make_hps("Lisensi Microsoft 365 E3", 2026,
    items=[
        ("Lisensi E3 user (2500 user, 12 bulan)", "User-Bulan", 30000, 110_000),
        ("Setup tenant & migration", "Lump sum", 1, 100_000_000),
    ],
    vendors=[
        ("PT Mitra Integrasi Informatika", 3_400_000_000),
        ("PT Sinergi Wahana Gemilang", 3_510_000_000),
    ],
    filename="HPS-Lisensi-M365-2026.pdf")

print("\n[3] RFI (Request for Information) — 5 file")
make_rfi("PT Cyber Network Indonesia (CNI)", "Jl. TB Simatupang Kav. 88, Jakarta Selatan",
         "Layanan Cloud PSrE Induk", 2026, 6_017_600_000, "RFI-CNI-Cloud-PSrE-2026.pdf")
make_rfi("PT Telkom Sigma", "Wisma Telkom, Jl. Gatot Subroto, Jakarta",
         "Layanan Cloud PSrE Induk", 2025,  # nb: salah TA
         None, "RFI-Telkom-Sigma-Cloud-2025.pdf",
         alasan_tolak="tim kami sedang berkomitmen penuh pada project migrasi data sektor publik lain.")
make_rfi("PT Biznet Networks", "MidPlaza 2, Jakarta Pusat",
         "Layanan Cloud PSrE Induk", 2026, 5_980_000_000, "RFI-Biznet-Cloud-PSrE-2026.pdf")
make_rfi("PT DCI Indonesia", "DCI JK1, Cibitung",
         "Data Center Tier-3", 2026, 8_450_000_000, "RFI-DCI-DataCenter-2026.pdf")
make_rfi("PT Mitra Integrasi Informatika", "Menara Astra Lt. 23, Jakarta",
         "Lisensi Microsoft 365", 2026, 3_400_000_000, "RFI-MII-Lisensi-M365-2026.pdf")

print("\n[4] KONTRAK — 2 file")
make_kontrak("HK.02.04/PSrE/05/2026", "Layanan Cloud PSrE Induk", 2026,
             "PT Cyber Network Indonesia", 5_576_851_733, "12 bulan",
             "KONTRAK-Cloud-PSrE-CNI-2026.pdf")
make_kontrak("HK.02.04/DC/06/2026", "Data Center & DRC Tier-3", 2026,
             "PT DCI Indonesia", 8_200_000_000, "24 bulan",
             "KONTRAK-DataCenter-DCI-2026.pdf")

print("\n[5] TOR (Term of Reference) — 4 file")
make_tor("Pengembangan Aplikasi Pemantauan Perlindungan Data Pribadi", 2026,
         "Aplikasi web monitoring kepatuhan UU PDP",
         "1 aplikasi siap operasi di Q4 2026",
         2_450_000_000, "TOR-App-PDP-2026.pdf")
make_tor("Survei Indeks Literasi Digital Nasional", 2026,
         "Laporan survei nasional dengan 34 provinsi",
         "Coverage minimum 5.000 responden",
         3_200_000_000, "TOR-Survei-Literasi-Digital-2026.pdf")
make_tor("Sosialisasi UU 27/2022 tentang PDP", 2026,
         "Materi & event sosialisasi di 10 kota",
         "10 event dengan minimum 200 peserta/event",
         1_800_000_000, "TOR-Sosialisasi-PDP-2026.pdf")
make_tor("Pengembangan Pusat Data Nasional Tahap II", 2026,
         "Infrastruktur PDN ekspansi 2x kapasitas",
         "Go-live Desember 2026",
         15_000_000_000, "TOR-PDN-Tahap-2-2026.pdf")

print("\n[6] RAB (Excel) — 3 file")
make_rab("Pengembangan Aplikasi Pemantauan PDP", 2026,
    items=[
        ("521211", "Belanja Bahan", 1, 50_000_000),
        ("521213", "Honorarium Output Kegiatan", 1, 350_000_000),
        ("521219", "Belanja Barang Non-Operasional Lainnya", 1, 100_000_000),
        ("522151", "Belanja Jasa Profesi (konsultan)", 1, 800_000_000),
        ("522191", "Belanja Jasa Lainnya (pengembangan)", 1, 950_000_000),
        ("524111", "Belanja Perjalanan Biasa", 1, 200_000_000),
    ],
    filename="RAB-App-PDP-2026.xlsx")
make_rab("Survei Indeks Literasi Digital Nasional", 2026,
    items=[
        ("521211", "Belanja Bahan kuesioner & alat", 1, 80_000_000),
        ("521213", "Honor enumerator (34 provinsi)", 34, 25_000_000),
        ("521219", "Belanja sosialisasi & FGD", 1, 220_000_000),
        ("524111", "Belanja Perjalanan Dinas Dalam Negeri", 1, 1_150_000_000),
        ("522151", "Belanja Jasa Profesi (akademisi)", 1, 500_000_000),
    ],
    filename="RAB-Survei-Literasi-2026.xlsx")
make_rab("Sosialisasi UU PDP", 2026,
    items=[
        ("521211", "Belanja Bahan (materi cetak, banner)", 1, 150_000_000),
        ("521213", "Honor narasumber & moderator", 10, 30_000_000),
        ("521219", "Belanja Konsumsi & ATK event", 10, 60_000_000),
        ("524111", "Belanja Perjalanan ke 10 kota", 1, 750_000_000),
    ],
    filename="RAB-Sosialisasi-PDP-2026.xlsx")

print("\n[7] Surat Tugas (ST) — 2 file")
make_st("51/IJ.3/KP.01.06/03/2026",
        "Reviu Pengadaan Layanan Cloud PSrE Induk TA 2026", 2026,
        tim=[
            ("Sarah Aulia, S.Kom.", "198501012010012001", "Anggota Tim"),
            ("Citra Lestari, SE.", "198803152012012002", "Anggota Tim"),
            ("Budi Hartono, SE., Ak.", "197505152005011001", "Ketua Tim"),
            ("Audi Wibowo, SE., M.Ak.", "197001012000011001", "Pengendali Teknis"),
        ],
        filename="ST-51-Reviu-Cloud-PSrE-2026.pdf")
make_st("78/IJ.3/KP.01.06/04/2026",
        "Reviu RKA-K/L Direktorat Pengendalian Aplikasi TA 2026", 2026,
        tim=[
            ("Sarah Aulia, S.Kom.", "198501012010012001", "Anggota Tim"),
            ("Budi Hartono, SE., Ak.", "197505152005011001", "Ketua Tim"),
            ("Audi Wibowo, SE., M.Ak.", "197001012000011001", "Pengendali Teknis"),
        ],
        filename="ST-78-Reviu-RKA-DIT-PENGENDALIAN-2026.pdf")

print("\n[8] Kartu Penugasan (KP) — 2 file")
make_kp("KP-156/IJ.3/2026", "51/IJ.3/KP.01.06/03/2026",
        "Reviu Pengadaan Layanan Cloud PSrE Induk TA 2026", 2026,
        filename="KP-156-Reviu-Cloud-PSrE-2026.pdf")
make_kp("KP-203/IJ.3/2026", "78/IJ.3/KP.01.06/04/2026",
        "Reviu RKA-K/L DIT Pengendalian TA 2026", 2026,
        filename="KP-203-Reviu-RKA-2026.pdf")

print("\n[9] PKP (Program Kerja Pengawasan) — 2 file")
make_pkp("REVIU PENGADAAN LAYANAN CLOUD PSrE", 2026,
    sasaran=[
        ("S-PBJ-01", "Kelengkapan & Kewajaran KAK",
         "Cek 7 blok KAK; verifikasi SLA & jadwal", "Sarah Aulia"),
        ("S-PBJ-02", "Kewajaran HPS & RFI",
         "Cek minimum 2 RFI valid (Perpres 16 Ps 26.5)", "Sarah Aulia"),
        ("S-PBJ-03", "Dasar Hukum HPS & Konsistensi TA",
         "Verifikasi Perpres 12/2021, SBM TA 2026", "Sarah Aulia"),
        ("S-PBJ-04", "Metode Pemilihan",
         "Cek konsistensi metode KAK ↔ Kontrak", "Sarah Aulia"),
    ],
    filename="PKP-Reviu-Cloud-PSrE-2026.pdf")

make_pkp("REVIU RKA-K/L DIT PENGENDALIAN APLIKASI", 2026,
    sasaran=[
        ("S-RKA-01", "Kelengkapan 7 blok substansi TOR",
         "Cek latar belakang, dasar hukum, dst", "Sarah Aulia"),
        ("S-RKA-02", "Kewajaran SBM/SBK per akun",
         "Bandingkan dengan PMK SBM TA 2026", "Sarah Aulia"),
        ("S-RKA-03", "Cascading anggaran",
         "Cek RAB konsisten dengan TOR & RO", "Sarah Aulia"),
    ],
    filename="PKP-Reviu-RKA-DIT-PENGENDALIAN-2026.pdf")

# ============================================================
# README
# ============================================================
readme = OUT / "README.md"
readme.write_text("""# Dummy Test Documents — Audit AI v7

Folder ini berisi dokumen dummy realistic untuk testing UI workflow di audit-system-v7.
Generated otomatis — semua konten fiktif, tidak terkait dengan transaksi nyata.

## Inventaris

### Reviu Pengadaan (skill: `reviu-pengadaan`)

**5 KAK** — Kerangka Acuan Kerja
- `KAK-Layanan-Cloud-PSrE-2026.pdf` — Cloud PSrE, Rp 5,5 M, SLA 99,9%
- `KAK-Data-Center-DRC-2026.pdf` — DC Tier-3, Rp 8,2 M
- `KAK-Pelatihan-Cyber-Security-2026.pdf` — Pelatihan, Rp 450 jt
- `KAK-App-SatuData-2026.pdf` — App Mobile, Rp 1,85 M
- `KAK-Lisensi-M365-2026.pdf` — Lisensi MS 365, Rp 3,4 M

**4 HPS** — Harga Perkiraan Sendiri (rincian item + tabel vendor RFI)
- `HPS-Cloud-PSrE-2026.pdf` ⚠️ **berisi anomali**: RFI Telkom Sigma menolak (1 valid + 3 invalid)
- `HPS-Data-Center-2026.pdf`
- `HPS-Pelatihan-Cyber-2026.pdf`
- `HPS-Lisensi-M365-2026.pdf`

**5 RFI** — Request for Information dari vendor
- `RFI-CNI-Cloud-PSrE-2026.pdf` — penawaran valid
- `RFI-Telkom-Sigma-Cloud-2025.pdf` ⚠️ **surat penolakan + TA salah (2025 vs 2026)**
- `RFI-Biznet-Cloud-PSrE-2026.pdf`
- `RFI-DCI-DataCenter-2026.pdf`
- `RFI-MII-Lisensi-M365-2026.pdf`

**2 KONTRAK**
- `KONTRAK-Cloud-PSrE-CNI-2026.pdf`
- `KONTRAK-DataCenter-DCI-2026.pdf`

### Reviu RKA-K/L (skill: `reviu-rka-kl`)

**4 TOR** — Term of Reference (struktur lengkap 7-8 blok)
- `TOR-App-PDP-2026.pdf` — Rp 2,45 M
- `TOR-Survei-Literasi-Digital-2026.pdf` — Rp 3,2 M
- `TOR-Sosialisasi-PDP-2026.pdf` — Rp 1,8 M
- `TOR-PDN-Tahap-2-2026.pdf` — Rp 15 M

**3 RAB** — Rencana Anggaran Biaya (Excel, tabel akun)
- `RAB-App-PDP-2026.xlsx`
- `RAB-Survei-Literasi-2026.xlsx`
- `RAB-Sosialisasi-PDP-2026.xlsx`

### Dokumen Umum

**2 ST** — Surat Tugas
- `ST-51-Reviu-Cloud-PSrE-2026.pdf` — penugasan reviu pengadaan
- `ST-78-Reviu-RKA-DIT-PENGENDALIAN-2026.pdf` — penugasan reviu RKA-K/L

**2 KP** — Kartu Penugasan
- `KP-156-Reviu-Cloud-PSrE-2026.pdf`
- `KP-203-Reviu-RKA-2026.pdf`

**2 PKP** — Program Kerja Pengawasan (sasaran + langkah kerja)
- `PKP-Reviu-Cloud-PSrE-2026.pdf` — 4 sasaran (S-PBJ-01 s.d. S-PBJ-04)
- `PKP-Reviu-RKA-DIT-PENGENDALIAN-2026.pdf` — 3 sasaran (S-RKA-01 s.d. S-RKA-03)

## Cara Pakai

### Skenario 1: Test Reviu Pengadaan

1. Buat penugasan baru dengan skill **Reviu Pengadaan**, obyek "Pengadaan Layanan Cloud PSrE Induk"
2. Upload ke tab Dokumen:
   - `ST-51-...pdf` (ST)
   - `KP-156-...pdf` (KP)
   - `PKP-Reviu-Cloud-PSrE-2026.pdf` (PKP)
   - `KAK-Layanan-Cloud-PSrE-2026.pdf` (KAK)
   - `HPS-Cloud-PSrE-2026.pdf` (HPS)
   - 4 file RFI (`RFI-CNI-...`, `RFI-Telkom-Sigma-...`, `RFI-Biznet-...`)
   - `KONTRAK-Cloud-PSrE-CNI-2026.pdf` (KONTRAK)
3. Ingestion auto-trigger, tunggu status semua `READY`
4. Edit `context.md` dan `_PKP/sasaran-assignment.json` (set sasaran S-PBJ-01..04 assigned ke "Sarah Aulia")
5. Tab **Chat AT** → jalankan agen
6. Expected temuan: anomali RFI Telkom Sigma (penolakan + TA salah), HPS hanya 1 sumber harga valid

### Skenario 2: Test Reviu RKA-K/L

Pakai `ST-78`, `KP-203`, `PKP-Reviu-RKA-...`, plus 1-2 TOR + RAB yang sesuai.

## Catatan

Generator script: `dummy-test-docs/_generator.py` (jika perlu regenerate).
Folder ini di-gitignore — tidak ter-commit ke repo.
""", encoding="utf-8")
print(f"\n  ✓ README.md")

# List all
print("\n" + "="*60)
print("SUMMARY")
print("="*60)
files = sorted(OUT.iterdir())
print(f"Total files: {len(files)}")
total_size = sum(f.stat().st_size for f in files if f.is_file())
print(f"Total size: {total_size:,} bytes ({total_size/1024/1024:.1f} MB)")

